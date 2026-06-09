import json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from datetime import datetime
import re

EXCEL = r'D:\工作\华Hunt队伍强度表\战力表260609.xlsx'
OUT_DIR = r'D:\工作\华Hunt队伍强度表\华hunt-ranking\data'

wb = openpyxl.load_workbook(EXCEL, data_only=True)

ws_power = wb['战力表']
teams = []
for r in range(2, 30):
    tier = ws_power.cell(row=r, column=1).value
    rank = ws_power.cell(row=r, column=2).value
    name = ws_power.cell(row=r, column=3).value
    score = ws_power.cell(row=r, column=4).value
    if not name:
        continue
    teams.append({
        'name': str(name).strip().replace('（解散）', ''),
        'tier': str(tier).strip() if tier else '',
        'rank': str(rank).strip() if rank else '',
        'score': round(score, 2) if isinstance(score, (int, float)) else score,
    })

ws_rec = wb['战绩表']

def normalize_date(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d']:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    m = re.match(r'(\d{4})/(\d{1,2})/(\d{1,2})', s)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'
    return s if len(s) >= 7 else s

competitions = []
for r in range(3, 19):
    b = ws_rec.cell(row=r, column=2)
    if not b.value:
        continue
    date_val = ws_rec.cell(row=r, column=3).value
    competitions.append({
        'id': str(b.value).strip(),
        'date': normalize_date(date_val)
    })

# Load URLs
links = {}
try:
    with open(r'D:\工作\华Hunt队伍强度表\链接.txt', 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line: continue
            p = line.split('\t')
            if len(p) >= 2: links[p[0].strip()] = p[1].strip()
except: pass
for c in competitions:
    c['url'] = links.get(c['id'], '')

name_mapping = {
    'simplicissimus': '4s', '灵境TES': '灵境TES',
    '识文解意的追迹者': '追迹者', '鸡': '鸡', '杂鱼': '杂鱼',
    'Anonymous': 'Anon-ymous', 'Reviver Team': 'Reviver',
    '解谜怪盗团': '怪盗团', '海拉鲁野炊组': '野炊组',
    '金牌线': '金牌线', '闲散解谜！': '闲散解谜',
    '喵喵喵': '喵喵喵', 'zitmen': 'zitmen',
    '谜途旅人': '谜途旅人', 'JUSTeam': 'JUST-eam',
    '幼稚园的小猫咪': '小猫咪', '世界美食official': '世界美食',
    'rusty lake rambler': 'rlr', '天翔战队ZOO连者！': '天翔战队',
    '看看题': '看看题', '反卷': '反卷', '何以为我': '何以为我',
    '凑不出': '凑不出', '深色墨点': '深色墨点', '绿色圆圈': '绿色圆圈',
    'Puzzlesweeper': 'Puzzle-sweeper', 'Not A Hunter': 'NAH',
    'Cyfur': 'Cyfur',
}

team_header_map = {}
for col in range(5, ws_rec.max_column + 1, 2):
    h = ws_rec.cell(row=2, column=col)
    if h.value:
        team_header_map[str(h.value).strip()] = col

rankings = {}
for team in teams:
    tn = team['name']
    abbr = name_mapping.get(tn, tn)
    rankings[tn] = {}
    if abbr not in team_header_map:
        continue
    rank_col = team_header_map[abbr]
    for i, comp in enumerate(competitions):
        row_num = 3 + i
        rv = ws_rec.cell(row=row_num, column=rank_col).value
        if rv is not None:
            try:
                rankings[tn][comp['id']] = int(rv)
            except (ValueError, TypeError):
                s = str(rv).strip()
                rankings[tn][comp['id']] = s if s else None
        else:
            rankings[tn][comp['id']] = None

# ===== Competition results (top teams with members) =====
skip = {'战力表', '战绩表', '二大赛事前排', '追迹者战绩', '非人答案'}
comp_results = {}
for sn in wb.sheetnames:
    if sn in skip:
        continue
    ws = wb[sn]
    results = []
    time_col = None
    member_cols = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        if h and '用时' in str(h):
            time_col = col
        elif h and '队员' in str(h) and '生数据' not in str(h):
            member_cols.append(col)

    for r in range(2, ws.max_row + 1):
        rank = ws.cell(row=r, column=1).value
        if not rank:
            continue
        raw_name = ws.cell(row=r, column=2).value
        clean_name = ws.cell(row=r, column=3).value
        time_val = ''
        if time_col:
            tv = ws.cell(row=r, column=time_col).value
            time_val = str(tv).strip() if tv else ''
        members = []
        for mc in member_cols:
            mv = ws.cell(row=r, column=mc).value
            if mv:
                members.append(str(mv).strip())
        results.append({
            'rank': int(rank) if isinstance(rank, (int, float)) else str(rank),
            'rawName': str(raw_name).strip() if raw_name else '',
            'cleanName': str(clean_name).strip() if clean_name else '',
            'time': time_val,
            'members': members,
        })
    comp_results[sn] = results

with open(f'{OUT_DIR}/teams.json', 'w', encoding='utf-8') as f:
    json.dump(teams, f, ensure_ascii=False, indent=2)
with open(f'{OUT_DIR}/competitions.json', 'w', encoding='utf-8') as f:
    json.dump(competitions, f, ensure_ascii=False, indent=2)
with open(f'{OUT_DIR}/rankings.json', 'w', encoding='utf-8') as f:
    json.dump(rankings, f, ensure_ascii=False, indent=2)
with open(f'{OUT_DIR}/comp_results.json', 'w', encoding='utf-8') as f:
    json.dump(comp_results, f, ensure_ascii=False, indent=2)

print(f'Exported: {len(teams)} teams, {len(competitions)} competitions, {len(comp_results)} comp sheets')
